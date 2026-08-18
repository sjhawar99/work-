"""Reading sheets into an addressable grid.

The workbook is opened **read-only**. Nothing in this package ever writes to it: the
canonical source is the Google Sheet, and `input/workbook.xlsx` is an export artifact
(AGENTS.md rule 6).
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from apex_ads.ingest.errors import MissingSheetError
from apex_ads.util.hashing import sha256_file
from apex_ads.util.text import is_null_token, normalise_text

Cell = object


@dataclass(frozen=True)
class Row:
    """One worksheet row: its 1-based number and its cell values."""

    number: int
    cells: tuple[Cell, ...]

    def value(self, index: int) -> Cell:
        return self.cells[index] if 0 <= index < len(self.cells) else None

    def text(self, index: int) -> str:
        value = self.value(index)
        return "" if value is None else normalise_text(str(value))

    @property
    def is_blank(self) -> bool:
        return all(is_null_token(cell) or cell is None for cell in self.cells)

    @property
    def populated_count(self) -> int:
        return sum(1 for cell in self.cells if cell is not None and str(cell).strip())


@dataclass(frozen=True)
class Sheet:
    """One worksheet, as rows."""

    name: str
    rows: tuple[Row, ...]


@dataclass(frozen=True)
class Workbook:
    """A whole workbook plus the provenance needed to prove which file it was."""

    path: Path
    sha256: str
    mtime: datetime
    sheets: dict[str, Sheet]

    def sheet(self, name: str) -> Sheet:
        if name not in self.sheets:
            raise MissingSheetError(name, sorted(self.sheets))
        return self.sheets[name]


def load(path: Path) -> Workbook:
    """Read every sheet into memory, read-only, recording the file's hash and mtime."""
    if not path.is_file():
        raise FileNotFoundError(f"workbook not found: {path}")

    book = load_workbook(path, data_only=True, read_only=True)
    try:
        sheets: dict[str, Sheet] = {}
        for name in book.sheetnames:
            worksheet = book[name]
            rows = tuple(
                Row(number=index, cells=tuple(values))
                for index, values in enumerate(worksheet.iter_rows(values_only=True), start=1)
            )
            sheets[name] = Sheet(name=name, rows=rows)
    finally:
        book.close()

    stat = path.stat()
    return Workbook(
        path=path,
        sha256=sha256_file(path),
        mtime=datetime.fromtimestamp(stat.st_mtime),
        sheets=sheets,
    )
